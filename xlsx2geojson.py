import argparse
import json
import sys
from pathlib import Path

from datetime import datetime, timezone
from typing import Optional, Any

from openpyxl.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.reader.excel import load_workbook
from validataclass.exceptions import ValidationError

sys.path.append(str(Path(Path(__file__).parent.parent, "parkapi-sources-v3/src")))

from parkapi_sources.exceptions import (
    ImportParkingSiteException,
    ImportParkingSpotException,
)
from parkapi_sources.models import StaticParkingSiteInput, StaticParkingSpotInput

from parkapi_sources.converters.base_converter import (
    ParkingSiteBaseConverter,
    ParkingSpotBaseConverter,
)
from parkapi_sources.converters.base_converter.push import (
    NormalizedXlsxConverter,
    XlsxConverter,
)
from parkapi_sources.models import SourceInfo, ExcelOpeningTimeInput
from parkapi_sources.util import ConfigHelper, RequestHelper
from decimal import Decimal
from validataclass.dataclasses import Default, validataclass
from validataclass.validators import DataclassValidator
from parkapi_sources.validators import (
    ExcelNoneable,
    GermanDurationIntegerValidator,
    NumberCastingStringValidator,
)

import warnings


parser = argparse.ArgumentParser(
    prog="Xlsx2Geojson Conversion Script",
    description="This script helps to convert Excel Reference Table to ParkAPI Geojson",
)

parser.add_argument("source_uid")
parser.add_argument("source_group")
args = parser.parse_args()
source_uid: str = args.source_uid
source_group: str = args.source_group

if not source_group:
    sys.exit("Error: please add a source_uid as first argument")

if not source_group:
    sys.exit("Error: please add a source_group e.g. parking-spots, parking-sites")

if source_group == "parking-sites":
    file_path: Path = Path(f"sources/{str(source_uid)}.xlsx")
elif source_group == "parking-spots":
    file_path: Path = Path(f"sources/parking-spots/{str(source_uid)}.xlsx")

if not file_path.is_file():
    sys.exit(f"Error: please add an Excel file with name '{file_path}'")


def serialize(obj):
    if isinstance(obj, (bool, int, float)):
        return obj
    elif isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, datetime):
        return obj.isoformat()
    elif hasattr(obj, "value"):  # Check for Enums
        return obj.value
    elif isinstance(obj, list):
        return [
            serialize(item)
            if not isinstance(item, dict)
            else {k: serialize(v) for k, v in item.items() if v is not None}
            for item in obj
        ]
    elif obj.__class__.__name__ == "UnsetValueType":  # Check for UnsetValue
        return None
    return str(obj)


def filter_none(data: dict) -> dict:
    return {key: serialize(value) for key, value in data.items() if value is not None}


def to_single_line(text: str) -> str:
    return " ".join(text.strip().splitlines())


def normalize_text(text: str) -> str:
    if isinstance(text, str):
        return text.strip().lower()
    return text


@validataclass
class ExcelStaticParkingSpotInput(StaticParkingSpotInput):
    uid: str = NumberCastingStringValidator(min_length=1, max_length=256)
    max_stay: Optional[int] = (
        ExcelNoneable(GermanDurationIntegerValidator()),
        Default(None),
    )


class HeaderMappingMixin:
    # For additional attributes not in Generic Converter and if column names changes
    additional_header_rows: dict[str, str] = {
        "Einfahrtshöhe (cm)": "max_height",
        "Zweck der Anlage": "purpose",
        "Überwacht?": "supervision_type",
        "Anzahl Stellplätze Carsharing": "capacity_carsharing",
        "Anzahl Stellplätze Lademöglichkeit": "capacity_charging",
        "Anzahl Stellplätze Frauen": "capacity_woman",
        "Anzahl Stellplätze Behinderte": "capacity_disabled",
        "Anzahl Stellplätze Familien": "capacity_family",
        "Anzahl Stellplätze Bus": "capacity_bus",
        "Anzahl Stellplätze Lastwagen": "capacity_truck",
        "Park+Ride?": "park_and_ride_type",
        "Einfahrtshöhe": "max_height",
        "Überdacht?": "is_covered",
        "ID": "uid",
        "Name": "name",
        "Art der Anlage": "type",
        "Widmung": "restricted_to_type",
        "Längengrad": "lon",
        "Breitengrad": "lat",
        "Geometry": "geojson",
        "Maximale Parkdauer": "max_stay",
        "24/7 geöffnet?": "opening_hours_is_24_7",
        "Öffnungszeiten Mo-Fr Beginn": "opening_hours_weekday_begin",
        "Öffnungszeiten Mo-Fr Ende": "opening_hours_weekday_end",
        "Öffnungszeiten Sa Beginn": "opening_hours_saturday_begin",
        "Öffnungszeiten Sa Ende": "opening_hours_saturday_end",
        "Öffnungszeiten So Beginn": "opening_hours_sunday_begin",
        "Öffnungszeiten So Ende": "opening_hours_sunday_end",
        "Gebührenpflichtig?": "has_fee",
        "Adresse - Straße und Nummer": "street_number",
        "Adresse - PLZ und Stadt": "postcode_city",
    }

    def get_mapping_by_header(
        self, row: tuple[Cell, ...], expected_header_row: dict[str, str]
    ) -> dict[str, int]:
        row_values = [
            str(cell.value).strip().replace("\n", "") if cell.value else ""
            for cell in row
        ]
        normalized_headers = {
            header.lower(): idx for idx, header in enumerate(row_values)
        }

        mapping: dict[str, int] = {}
        for expected_header, target_field in expected_header_row.items():
            expected_normalized = expected_header.lower().strip()
            if expected_normalized in normalized_headers:
                mapping[target_field] = normalized_headers[expected_normalized]
            else:
                warnings.warn(
                    f"⚠️ Missing header: '{expected_header}' for field '{target_field}'",
                    ImportWarning,
                )
        return mapping


class EnumTypeMappingMixin:
    type_mapping: dict[str, str] = {
        "parkplatz": "OFF_STREET_PARKING_GROUND",
        "parkhaus": "CAR_PARK",
        "tiefgarage": "UNDERGROUND",
    }

    purpose_type_mapping: dict[str, str] = {
        "auto": "CAR",
        "fahrrad": "BIKE",
    }

    supervision_type_mapping: dict[str, str] = {
        True: "YES",
        False: "NO",
        "video": "VIDEO",
        "ja": "YES",
        "nein": "NO",
        "bewacht": "ATTENDED",
    }

    restricted_to_type_mapping: dict[str, str] = {
        "ladesäule": "CHARGING",
        "familie": "FAMILY",
        "handicap": "DISABLED",
    }

    park_and_ride_type_mapping: dict[str, str] = {
        "fahrgemeinschaft": "CARPOOL",
        "bahn": "TRAIN",
        "bus": "BUS",
        "straßenbahn": "TRAM",
        "ja": "YES",
        "nein": "NO",
    }


class Xlsx2GeojsonParkingSites(
    HeaderMappingMixin,
    EnumTypeMappingMixin,
    NormalizedXlsxConverter,
    ParkingSiteBaseConverter,
):
    source_info = SourceInfo(
        uid=source_uid,
        name="Convert ParkingSites Reference Table to Geojson",
        has_realtime_data=True,
    )

    def __init__(self, config: Optional[dict] = None):
        self.config = config or {}
        self.config_helper = ConfigHelper(config=self.config)
        self.request_helper = RequestHelper(config_helper=self.config_helper)
        super().__init__(self.config_helper, self.request_helper)

        self.header_row = {
            **{
                key: value
                for key, value in super().header_row.items()
                if value not in self.additional_header_rows.values()
            },
            **self.additional_header_rows,
        }

    def map_row_to_parking_site_dict(
        self,
        mapping: dict[str, int],
        row: tuple[Cell, ...],
        **kwargs: Any,
    ) -> dict[str, Any]:
        parking_site_dict = super().map_row_to_parking_site_dict(mapping, row, **kwargs)

        for field in mapping.keys():
            parking_site_dict[field] = (
                row[mapping[field]].value.strip()
                if isinstance(row[mapping[field]].value, str)
                else row[mapping[field]].value
            )

        parking_site_dict["max_height"] = (
            round(parking_site_dict.get("max_height"))*100
            if isinstance(parking_site_dict.get("max_height"), float)
            else parking_site_dict.get("max_height")*100
        )
        parking_site_dict["max_stay"] = (
            round(parking_site_dict.get("max_stay"))
            if isinstance(parking_site_dict.get("max_stay"), float)
            else parking_site_dict.get("max_stay")
        )
        parking_site_dict["opening_hours"] = parking_site_dict["opening_hours"].replace(
            "00:00-00:00", "00:00-24:00"
        )
        parking_site_dict["fee_description"] = (
            to_single_line(str(parking_site_dict["fee_description"]))
            if parking_site_dict["fee_description"]
            else None
        )
        parking_site_dict["purpose"] = self.purpose_type_mapping.get(
            normalize_text(parking_site_dict.get("purpose"))
        )
        parking_site_dict["type"] = self.type_mapping.get(
            normalize_text(parking_site_dict.get("type")), "OFF_STREET_PARKING_GROUND"
        )
        parking_site_dict["supervision_type"] = self.supervision_type_mapping.get(
            normalize_text(parking_site_dict.get("supervision_type")),
        )
        parking_site_dict["park_and_ride_type"] = self.park_and_ride_type_mapping.get(
            normalize_text(parking_site_dict.get("park_and_ride_type")),
        )
        if (
            "street_number" in parking_site_dict
            and "postcode_city" in parking_site_dict
        ):
            parking_site_dict["address"] = (
                f"""{parking_site_dict["street_number"]}, {parking_site_dict["postcode_city"]}"""
            )

        parking_site_dict["static_data_updated_at"] = datetime.now(
            tz=timezone.utc
        ).isoformat()

        parking_site_dict["lat"] = f"{parking_site_dict['lat']}"
        parking_site_dict["lon"] = f"{parking_site_dict['lon']}"
        parking_site_dict["has_realtime_data"] = self.source_info.has_realtime_data

        return parking_site_dict

    def handle_xlsx(
        self, workbook: Workbook
    ) -> tuple[list[StaticParkingSiteInput], list[ImportParkingSiteException]]:
        worksheet = workbook.active
        mapping: dict[str, int] = self.get_mapping_by_header(
            next(worksheet.rows), self.header_row
        )

        static_parking_site_errors: list[ImportParkingSiteException] = []
        static_parking_site_inputs: list[StaticParkingSiteInput] = []

        for row in worksheet.iter_rows(min_row=2):
            # ignore empty lines as LibreOffice sometimes adds empty rows at the end of a file
            if row[0].value is None:
                continue
            parking_site_dict = self.map_row_to_parking_site_dict(
                mapping=mapping,
                row=row,
                column_names=[cell.value for cell in next(worksheet.rows)],
            )

            try:
                static_parking_site_input = filter_none(
                    self.static_parking_site_validator.validate(
                        parking_site_dict
                    ).to_dict()
                )
                static_parking_site_inputs.append(
                    {
                        "type": "Feature",
                        "properties": static_parking_site_input,
                        "geometry": {
                            "type": "Point",
                            "coordinates": [
                                static_parking_site_input["lon"],
                                static_parking_site_input["lat"],
                            ],
                        },
                    }
                )

            except ValidationError as e:
                static_parking_site_errors.append(
                    ImportParkingSiteException(
                        source_uid=self.source_info.uid,
                        parking_site_uid=parking_site_dict.get("uid"),
                        message=f"invalid static parking site data {parking_site_dict}: {e.to_dict()}",
                    )
                )
                continue

        return static_parking_site_inputs, static_parking_site_errors


class Xlsx2GeojsonParkingSpots(
    HeaderMappingMixin, EnumTypeMappingMixin, XlsxConverter, ParkingSpotBaseConverter
):
    source_info = SourceInfo(
        uid=source_uid,
        name="Convert ParkingSpots Reference Table to Geojson",
        has_realtime_data=True,
    )

    static_parking_spot_validator = DataclassValidator(ExcelStaticParkingSpotInput)
    excel_opening_time_validator = DataclassValidator(ExcelOpeningTimeInput)

    def __init__(self, config: Optional[dict] = None):
        self.config = config or {}
        self.config_helper = ConfigHelper(config=self.config)
        self.request_helper = RequestHelper(config_helper=self.config_helper)
        super().__init__(self.config_helper, self.request_helper)

    def handle_xlsx(
        self, workbook: Workbook
    ) -> tuple[list[StaticParkingSpotInput], list[ImportParkingSpotException]]:
        worksheet = workbook.active
        mapping: dict[str, int] = self.get_mapping_by_header(
            next(worksheet.rows), self.additional_header_rows
        )

        static_parking_spot_errors: list[ImportParkingSpotException] = []
        static_parking_spot_features: list[StaticParkingSpotInput] = []

        for row in worksheet.iter_rows(min_row=2):
            # ignore empty lines as LibreOffice sometimes adds empty rows at the end of a file
            if row[0].value is None:
                continue
            parking_spot_dict = self.map_row_to_parking_spot_dict(
                mapping=mapping,
                row=row,
                column_names=[cell.value for cell in next(worksheet.rows)],
            )

            try:
                static_parking_spot_input = filter_none(
                    self.static_parking_spot_validator.validate(
                        parking_spot_dict
                    ).to_dict()
                )

                static_parking_spot_features.append(
                    {
                        "type": "Feature",
                        "properties": static_parking_spot_input,
                        "geometry": {
                            "type": "Point",
                            "coordinates": [
                                static_parking_spot_input["lon"],
                                static_parking_spot_input["lat"],
                            ],
                        },
                    }
                )

            except ValidationError as e:
                static_parking_spot_errors.append(
                    ImportParkingSpotException(
                        source_uid=self.source_info.uid,
                        parking_spot_uid=parking_spot_dict.get("uid"),
                        message=f"invalid static parking spot data {parking_spot_dict}: {e.to_dict()}",
                    )
                )
                continue

        return static_parking_spot_features, static_parking_spot_errors

    def map_row_to_parking_spot_dict(
        self,
        mapping: dict[str, int],
        row: tuple[Cell, ...],
        column_names: list[str],
    ) -> dict[str, Any]:
        parking_spot_raw_dict: dict[str, str] = {}
        for field in mapping.keys():
            parking_spot_raw_dict[field] = (
                row[mapping[field]].value.strip()
                if isinstance(row[mapping[field]].value, str)
                else row[mapping[field]].value
            )

        parking_spot_dict = {
            key: value
            for key, value in parking_spot_raw_dict.items()
            if not key.startswith("opening_hours_")
        }
        opening_hours_input = self.excel_opening_time_validator.validate(
            {
                key: value
                for key, value in parking_spot_raw_dict.items()
                if key.startswith("opening_hours_")
            }
        )

        restricted_to_raw = parking_spot_raw_dict.get("restricted_to_type", "")
        restricted_to_type = (
            self.restricted_to_type_mapping.get(restricted_to_raw.lower().strip())
            if isinstance(restricted_to_raw, str)
            else None
        )
        restricted_to = {
            "type": restricted_to_type,
            "hours": opening_hours_input.get_osm_opening_hours(),
        }

        raw_type = normalize_text(parking_spot_dict.get("type"))
        parking_spot_dict["type"] = (
            self.type_mapping[raw_type.strip()]
            if isinstance(raw_type, str) and raw_type.strip() in self.type_mapping
            else None
        )

        parking_spot_dict["restrictions"] = [restricted_to]
        parking_spot_dict["has_realtime_data"] = self.source_info.has_realtime_data
        parking_spot_dict["purpose"] = self.purpose_type_mapping.get(
            normalize_text(parking_spot_dict.get("purpose"))
        )
        parking_spot_dict["static_data_updated_at"] = datetime.now(
            tz=timezone.utc
        ).isoformat()

        parking_spot_dict["lat"] = f"{parking_spot_dict['lat']}"
        parking_spot_dict["lon"] = f"{parking_spot_dict['lon']}"

        return parking_spot_dict


workbook = load_workbook(filename=str(file_path.absolute()))
static_parking_inputs, import_parking_exceptions = [], []

if source_group == "parking-sites":
    xlsx2geojson = Xlsx2GeojsonParkingSites()
elif source_group == "parking-spots":
    xlsx2geojson = Xlsx2GeojsonParkingSpots()

static_parking_inputs, import_parking_exceptions = xlsx2geojson.handle_xlsx(workbook)
static_geojson_parking_inputs: dict[str, Any] = {
    "type": "FeatureCollection",
    "features": static_parking_inputs,
}

geojson_file = file_path.with_suffix(".geojson")
with geojson_file.open("w") as file:
    json.dump(static_geojson_parking_inputs, file, ensure_ascii=False, indent=4)

if len(import_parking_exceptions) > 0:
    print(import_parking_exceptions)
print(
    f"Successful with {len(static_parking_inputs)} {source_group} and {len(import_parking_exceptions)} Errors"
)
